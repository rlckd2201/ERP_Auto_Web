from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import BankTransfer, VoucherLine, VoucherPayload
from .settings import ManagerProfile


VENDOR_ALIASES = {"업체명", "거래처명", "상호", "vendor", "vendor_name", "name"}
VENDOR_CODE_ALIASES = {"업체코드", "거래처코드", "코드", "vendor_code", "code"}
AMOUNT_ALIASES = {"금액", "지급액", "결제금액", "합계", "총액", "amount", "payment_amount"}


@dataclass(frozen=True)
class ColumnMap:
    vendor_name: int
    vendor_code: int | None
    amount: int
    labels: dict[str, str]


def _clean_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s\-_()/\\[\].:]+", "", text)


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _to_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip()
    if not text:
        return 0
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return 0
    try:
        amount = int(round(float(text)))
    except ValueError:
        return 0
    return -amount if negative else amount


def _parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("회계일은 YYYY-MM-DD 형식이어야 합니다.") from exc


def _previous_month_label(accounting_date: str) -> str:
    parsed = _parse_date(accounting_date)
    month = 12 if parsed.month == 1 else parsed.month - 1
    return f"{month}월"


def _find_column(headers: list[Any], aliases: set[str]) -> int | None:
    cleaned_aliases = {_clean_header(alias) for alias in aliases}
    for index, value in enumerate(headers):
        if _clean_header(value) in cleaned_aliases:
            return index
    return None


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, ColumnMap]:
    for row_index, row in enumerate(rows[:20], start=1):
        vendor_index = _find_column(list(row), VENDOR_ALIASES)
        amount_index = _find_column(list(row), AMOUNT_ALIASES)
        if vendor_index is None or amount_index is None:
            continue
        code_index = _find_column(list(row), VENDOR_CODE_ALIASES)
        labels = {
            "vendor_name": str(row[vendor_index] or ""),
            "amount": str(row[amount_index] or ""),
        }
        if code_index is not None:
            labels["vendor_code"] = str(row[code_index] or "")
        return row_index, ColumnMap(vendor_index, code_index, amount_index, labels)
    raise ValueError("엑셀에서 업체명과 금액 컬럼을 찾지 못했습니다.")


def _vendor_summary(month_label: str, vendor_name: str, vendor_code: str) -> str:
    vendor = vendor_name.strip() or "업체명 미확인"
    code = vendor_code.strip()
    suffix = f"{vendor}({code})" if code else vendor
    return f"{month_label} 수시결제 - {suffix}"


def _erp_clipboard_rows(lines: list[VoucherLine]) -> list[str]:
    rows: list[str] = []
    for line in lines:
        debit = line.amount if line.side == "debit" else 0
        credit = line.amount if line.side == "credit" else 0
        rows.append(f"{line.account_name}\t{debit}\t{credit}\t\t{line.summary}")
    return rows


def _erp_line_management_items(lines: list[VoucherLine]) -> list[dict[str, str]]:
    return [dict(line.management_items) for line in lines]


def _payable_management_items(vendor_code: str, vendor_name: str) -> dict[str, str]:
    return {"거래처": vendor_code.strip() or vendor_name.strip()}


def _bank_management_items(manager: ManagerProfile) -> dict[str, str]:
    return {
        "계좌번호": manager.bank_account_no,
        "금융기관지점": manager.bank_branch_name,
        "거래처": manager.bank_vendor,
    }


def _parse_header_total(value: Any) -> tuple[int, int]:
    text = str(value or "")
    numbers = [int(part.replace(",", "")) for part in re.findall(r"\d[\d,]*", text)]
    if not numbers:
        return 0, 0
    total = max(numbers)
    count = numbers[-1] if len(numbers) >= 2 else 0
    return total, count


def _read_bank_transfers(workbook: Any, max_rows: int = 800) -> list[BankTransfer]:
    if "인터넷뱅킹" not in workbook.sheetnames:
        return []
    sheet = workbook["인터넷뱅킹"]
    transfers: list[BankTransfer] = []
    blank_run = 0
    for row_index in range(1, max_rows + 1):
        bank_code = _clean_text(sheet.cell(row_index, 1).value)
        account_no = _clean_text(sheet.cell(row_index, 2).value)
        depositor = _clean_text(sheet.cell(row_index, 3).value)
        amount = _to_int(sheet.cell(row_index, 4).value)
        company_name = _clean_text(sheet.cell(row_index, 6).value)
        if amount > 0 and (account_no or depositor):
            blank_run = 0
            transfers.append(
                BankTransfer(
                    seq=len(transfers) + 1,
                    bank_code=bank_code,
                    account_no=account_no,
                    depositor=depositor,
                    amount=amount,
                    company_name=company_name,
                    source_row=row_index,
                )
            )
            continue
        blank_run += 1
        if transfers and blank_run >= 30:
            break
    return transfers


def _looks_like_sequence(amounts: list[int]) -> bool:
    if len(amounts) < 3:
        return False
    expected = list(range(1, len(amounts) + 1))
    return amounts == expected


def _cash_amount_source(
    cash_amounts: list[int],
    bank_transfers: list[BankTransfer],
    header_total: int,
) -> tuple[str, list[str]]:
    warnings: list[str] = []
    cash_total = sum(cash_amounts)
    bank_total = sum(transfer.amount for transfer in bank_transfers)
    if header_total and cash_total == header_total:
        return "cash", warnings
    if not header_total and cash_total > 0 and not _looks_like_sequence(cash_amounts):
        return "cash", warnings
    if bank_transfers and len(bank_transfers) == len(cash_amounts):
        if not header_total or bank_total == header_total:
            warnings.append("수시결제현금 H열이 승인서 합계와 맞지 않아 인터넷뱅킹 D열 금액을 같은 순번으로 보조 사용했습니다.")
            return "bank_fallback", warnings
    if header_total and cash_total != header_total:
        raise ValueError(
            f"수시결제현금 H열 합계({cash_total:,}원)가 승인서 합계({header_total:,}원)와 다릅니다. "
            "금액 열을 확인해 주세요."
        )
    return "cash", warnings


def _build_daeseung_cash_payload(
    workbook: Any,
    *,
    accounting_date: str,
    requester: str,
    source_filename: str,
    manager: ManagerProfile,
) -> VoucherPayload:
    cash_sheet = workbook["수시결제현금"]
    month_label = _previous_month_label(accounting_date)
    bank_transfers = _read_bank_transfers(workbook)
    header_total, header_count = _parse_header_total(cash_sheet.cell(6, 1).value)
    warnings: list[str] = []
    cash_rows: list[dict[str, Any]] = []
    blank_run = 0

    for row_index in range(9, 900):
        department = _clean_text(cash_sheet.cell(row_index, 4).value)
        vendor_code = _clean_text(cash_sheet.cell(row_index, 5).value)
        vendor_name = _clean_text(cash_sheet.cell(row_index, 6).value)
        original_summary = _clean_text(cash_sheet.cell(row_index, 7).value)
        payment_date = _clean_text(cash_sheet.cell(row_index, 3).value)
        cash_amount = _to_int(cash_sheet.cell(row_index, 8).value)
        if vendor_code and vendor_name:
            blank_run = 0
            cash_rows.append(
                {
                    "source_row": row_index,
                    "department": department,
                    "vendor_code": vendor_code,
                    "vendor_name": vendor_name,
                    "original_summary": original_summary,
                    "payment_date": payment_date,
                    "cash_amount": cash_amount,
                }
            )
            continue
        blank_run += 1
        if cash_rows and blank_run >= 30:
            break

    if not cash_rows:
        raise ValueError("수시결제현금 시트에서 전표 대상 행을 찾지 못했습니다.")
    cash_amounts = [int(row.get("cash_amount") or 0) for row in cash_rows]
    amount_source, amount_warnings = _cash_amount_source(cash_amounts, bank_transfers, header_total)
    warnings.extend(amount_warnings)
    if amount_source == "bank_fallback" and len(bank_transfers) != len(cash_rows):
        warnings.append(f"수시결제현금 {len(cash_rows)}행과 인터넷뱅킹 {len(bank_transfers)}행 수가 다릅니다.")

    lines: list[VoucherLine] = []
    for index, cash_row in enumerate(cash_rows):
        bank = bank_transfers[index] if amount_source == "bank_fallback" and index < len(bank_transfers) else None
        amount = bank.amount if bank else int(cash_row.get("cash_amount") or 0)
        if amount <= 0:
            warnings.append(f"{cash_row['source_row']}행 금액이 0 이하라 제외했습니다.")
            continue
        lines.append(
            VoucherLine(
                seq=len(lines) + 1,
                side="debit",
                account_name="미지급금(원화)",
                amount=amount,
                summary=_vendor_summary(month_label, cash_row["vendor_name"], cash_row["vendor_code"]),
                management_items=_payable_management_items(cash_row["vendor_code"], cash_row["vendor_name"]),
                vendor_name=cash_row["vendor_name"],
                vendor_code=cash_row["vendor_code"],
                department=cash_row["department"],
                original_summary=cash_row["original_summary"],
                payment_date=cash_row["payment_date"],
                source_sheet="수시결제현금",
                source_row=cash_row["source_row"],
            )
        )

    if not lines:
        raise ValueError("전표로 변환할 금액 행이 없습니다.")

    debit_total = sum(line.amount for line in lines)
    if header_total and header_total != debit_total:
        warnings.append(f"승인서 합계 {header_total:,}원과 전표 차변 합계 {debit_total:,}원이 다릅니다.")
    if header_count and header_count != len(lines):
        warnings.append(f"승인서 건수 {header_count}건과 전표 행 수 {len(lines)}건이 다릅니다.")

    lines.append(
        VoucherLine(
            seq=len(lines) + 1,
            side="credit",
            account_name=manager.bank_account_name,
            amount=debit_total,
            summary=f"{month_label} 수시결제 - {manager.bank_summary_name}",
            management_items=_bank_management_items(manager),
            source_sheet="generated",
        )
    )

    return VoucherPayload(
        accounting_date=accounting_date,
        voucher_month_label=month_label,
        company_key=manager.key,
        company_name=manager.company_name,
        erp_site_name=manager.erp_site_name,
        requester=requester,
        source_filename=source_filename,
        source_format="daeseung_cash_sheet_v1",
        source_row_count=len(cash_rows),
        header_total=header_total,
        debit_total=debit_total,
        credit_total=debit_total,
        line_count=len(lines),
        lines=lines,
        cash_processing_enabled=False,
        erp_line_management_items=_erp_line_management_items(lines),
        bank_transfers=bank_transfers if amount_source == "bank_fallback" else [],
        erp_clipboard_rows=_erp_clipboard_rows(lines),
        source_columns={
            "vendor_code": "수시결제현금!E",
            "vendor_name": "수시결제현금!F",
            "original_summary": "수시결제현금!G",
            "amount": "수시결제현금!H" if amount_source == "cash" else "인터넷뱅킹!D (수시결제현금 행 순번 매칭)",
        },
        warnings=warnings,
    )


def _build_generic_payload(
    workbook: Any,
    *,
    accounting_date: str,
    requester: str,
    source_filename: str,
    manager: ManagerProfile,
) -> VoucherPayload:
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("엑셀에 데이터가 없습니다.")

    header_row, columns = _find_header_row(rows)
    month_label = _previous_month_label(accounting_date)
    lines: list[VoucherLine] = []

    for source_row, row in enumerate(rows[header_row:], start=header_row + 1):
        vendor_name = str(row[columns.vendor_name] or "").strip() if len(row) > columns.vendor_name else ""
        vendor_code = ""
        if columns.vendor_code is not None and len(row) > columns.vendor_code:
            vendor_code = str(row[columns.vendor_code] or "").strip()
        amount = _to_int(row[columns.amount] if len(row) > columns.amount else None)
        if not vendor_name and not vendor_code and amount == 0:
            continue
        if amount <= 0:
            continue
        lines.append(
            VoucherLine(
                seq=len(lines) + 1,
                side="debit",
                account_name="미지급금(원화)",
                amount=amount,
                summary=_vendor_summary(month_label, vendor_name, vendor_code),
                management_items=_payable_management_items(vendor_code, vendor_name),
                vendor_name=vendor_name,
                vendor_code=vendor_code,
                source_sheet=sheet.title,
                source_row=source_row,
            )
        )

    if not lines:
        raise ValueError("전표로 변환할 금액 행이 없습니다.")

    debit_total = sum(line.amount for line in lines)
    lines.append(
        VoucherLine(
            seq=len(lines) + 1,
            side="credit",
            account_name=manager.bank_account_name,
            amount=debit_total,
            summary=f"{month_label} 수시결제 - {manager.bank_summary_name}",
            management_items=_bank_management_items(manager),
            source_sheet="generated",
        )
    )

    return VoucherPayload(
        accounting_date=accounting_date,
        voucher_month_label=month_label,
        company_key=manager.key,
        company_name=manager.company_name,
        erp_site_name=manager.erp_site_name,
        requester=requester,
        source_filename=source_filename,
        source_row_count=len(lines) - 1,
        debit_total=debit_total,
        credit_total=debit_total,
        line_count=len(lines),
        lines=lines,
        cash_processing_enabled=False,
        erp_line_management_items=_erp_line_management_items(lines),
        erp_clipboard_rows=_erp_clipboard_rows(lines),
        source_columns=columns.labels,
    )


def build_voucher_payload(
    path: Path,
    *,
    accounting_date: str,
    requester: str,
    source_filename: str,
    manager: ManagerProfile,
) -> VoucherPayload:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if "수시결제현금" in workbook.sheetnames:
            return _build_daeseung_cash_payload(
                workbook,
                accounting_date=accounting_date,
                requester=requester,
                source_filename=source_filename,
                manager=manager,
            )
        return _build_generic_payload(
            workbook,
            accounting_date=accounting_date,
            requester=requester,
            source_filename=source_filename,
            manager=manager,
        )
    finally:
        workbook.close()
