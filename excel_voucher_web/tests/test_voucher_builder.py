from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.accounts import AccountStore, protect_secret, unprotect_secret
from app.agent_adapter import (
    _apply_company_erp_credentials,
    _legacy_form_data,
    _resume_print_only_requested,
)
from app.settings import manager_profile, manager_profiles
from app.voucher_builder import build_voucher_payload


def _write_sample(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["업체명", "업체코드", "금액"])
    sheet.append(["가나다상사", "A001", 12000])
    sheet.append(["라마정밀", "B002", "3,000"])
    sheet.append(["금액없는행", "C003", 0])
    workbook.save(path)


def _write_cash_sheet_sample(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "수시결제현금"
    sheet["A6"] = "합 계  : 15,000원 2(件)"
    sheet.append([])
    sheet.append(["구분", "NO", "거래일자", "부서", "코드", "업체명", "적         요", "금액"])
    sheet.append(["", "", "", "", "", "", "", ""])
    sheet.append(["평택", 1, 531, "구매", "A001", "가나다상사", "소모품", 12000])
    sheet.append(["", 2, 531, "물류", "B002", "라마정밀", "운반비", 3000])
    workbook.save(path)


def _write_many_cash_sheet_sample(path: Path, count: int = 208) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "수시결제현금"
    total = sum(index * 1000 for index in range(1, count + 1))
    sheet["A6"] = f"합 계  : {total:,}원 {count}(件)"
    sheet.append([])
    sheet.append(["구분", "NO", "거래일자", "부서", "코드", "업체명", "적         요", "금액"])
    sheet.append(["", "", "", "", "", "", "", ""])
    for index in range(1, count + 1):
        code = f"HD{index:03d}"
        name = f"거래처{index:03d}"
        if index == count:
            code = "HD084"
            name = "현대리바트"
        sheet.append(["", index, 430, "구매", code, name, "수시결제", index * 1000])
    workbook.save(path)


def test_build_voucher_payload_adds_bank_credit_line(tmp_path: Path) -> None:
    source = tmp_path / "sample.xlsx"
    _write_sample(source)

    payload = build_voucher_payload(
        source,
        accounting_date="2026-06-20",
        requester="tester",
        source_filename="sample.xlsx",
        manager=manager_profile("daeseung"),
    )

    assert payload.debit_total == 15000
    assert payload.credit_total == 15000
    assert payload.line_count == 3
    assert payload.erp_site_name == "D1공장"
    assert payload.lines[0].side == "debit"
    assert payload.lines[0].account_name == "미지급금(원화)"
    assert payload.lines[0].summary == "5월 수시결제 - 가나다상사(A001)"
    assert payload.lines[0].management_items == {"거래처": "A001"}
    assert payload.lines[-1].side == "credit"
    assert payload.lines[-1].account_name == "보통예금"
    assert payload.lines[-1].summary == "5월 수시결제 - 신한은행"
    assert payload.lines[-1].management_items == {
        "계좌번호": "140-000-948562",
        "금융기관지점": "신한 수원금융센터",
        "거래처": "",
    }
    assert payload.erp_clipboard_rows[0].split("\t")[:5] == [
        "미지급금(원화)",
        "12000",
        "0",
        "",
        "5월 수시결제 - 가나다상사(A001)",
    ]
    assert payload.erp_clipboard_rows[-1].split("\t")[:5] == [
        "보통예금",
        "0",
        "15000",
        "",
        "5월 수시결제 - 신한은행",
    ]
    assert payload.erp_line_management_items[0] == {"거래처": "A001"}
    assert payload.erp_line_management_items[-1] == {
        "계좌번호": "140-000-948562",
        "금융기관지점": "신한 수원금융센터",
        "거래처": "",
    }
    assert payload.cash_processing_enabled is False

    legacy = _legacy_form_data(payload.model_dump(mode="json"))
    assert legacy["site_name"] == "D1공장"
    assert legacy["invoice_date"] == "2026-06-20"
    assert legacy["erp_row_count"] == payload.line_count
    assert legacy["cash_processing_enabled"] is False
    assert legacy["erp_line_management_items"][-1]["계좌번호"] == "140-000-948562"

    legacy_payload = payload.model_dump(mode="json")
    legacy_payload.pop("erp_line_management_items")
    for line in legacy_payload["lines"]:
        line.pop("management_items", None)
    legacy_from_lines = _legacy_form_data(legacy_payload)
    assert legacy_from_lines["erp_line_management_items"][0] == {"거래처": "A001"}
    assert legacy_from_lines["erp_line_management_items"][-1]["계좌번호"] == "140-000-948562"

    clipboard_only_payload = payload.model_dump(mode="json")
    clipboard_only_payload.pop("erp_line_management_items")
    clipboard_only_payload.pop("lines")
    legacy_from_clipboard = _legacy_form_data(clipboard_only_payload)
    assert legacy_from_clipboard["erp_line_management_items"][0] == {"거래처": "A001"}
    assert legacy_from_clipboard["erp_line_management_items"][1] == {"거래처": "B002"}
    assert legacy_from_clipboard["erp_line_management_items"][-1]["계좌번호"] == "140-000-948562"

    old_clipboard_payload = payload.model_dump(mode="json")
    old_clipboard_payload["erp_clipboard_rows"] = [
        row.replace("\t", "\t\t", 1) for row in old_clipboard_payload["erp_clipboard_rows"]
    ]
    legacy_from_old_clipboard = _legacy_form_data(old_clipboard_payload)
    assert legacy_from_old_clipboard["erp_clipboard_rows"][0].split("\t")[:5] == [
        "미지급금(원화)",
        "12000",
        "0",
        "",
        "5월 수시결제 - 가나다상사(A001)",
    ]


def test_build_voucher_payload_uses_cash_sheet_rows_only(tmp_path: Path) -> None:
    source = tmp_path / "cash.xlsx"
    _write_cash_sheet_sample(source)

    payload = build_voucher_payload(
        source,
        accounting_date="2026-06-20",
        requester="tester",
        source_filename="cash.xlsx",
        manager=manager_profile("daeseung"),
    )

    assert payload.source_format == "daeseung_cash_sheet_v1"
    assert payload.source_row_count == 2
    assert payload.debit_total == 15000
    assert payload.line_count == 3
    assert payload.lines[0].source_sheet == "수시결제현금"
    assert payload.lines[0].amount == 12000
    assert payload.lines[0].management_items == {"거래처": "A001"}
    assert payload.source_columns["amount"] == "수시결제현금!H"


def test_build_voucher_payload_keeps_last_vendor_before_bank_line(tmp_path: Path) -> None:
    source = tmp_path / "cash_many.xlsx"
    _write_many_cash_sheet_sample(source, count=208)

    payload = build_voucher_payload(
        source,
        accounting_date="2026-05-20",
        requester="tester",
        source_filename="cash_many.xlsx",
        manager=manager_profile("daeseung"),
    )

    assert payload.source_row_count == 208
    assert payload.line_count == 209
    assert payload.lines[-2].account_name == "미지급금(원화)"
    assert payload.lines[-2].vendor_code == "HD084"
    assert payload.lines[-2].management_items == {"거래처": "HD084"}
    assert payload.lines[-1].account_name == "보통예금"
    assert payload.lines[-1].management_items["계좌번호"] == "140-000-948562"
    assert payload.erp_line_management_items[207] == {"거래처": "HD084"}
    assert payload.erp_line_management_items[208]["계좌번호"] == "140-000-948562"
    assert payload.erp_clipboard_rows[207].split("\t")[0] == "미지급금(원화)"
    assert payload.erp_clipboard_rows[208].split("\t")[0] == "보통예금"


def test_only_daeseung_manager_is_enabled_for_upload() -> None:
    profiles = manager_profiles()
    assert profiles["daeseung"].enabled is True
    assert profiles["daeseung_precision"].enabled is False
    assert profiles["ilgwang"].enabled is False
    assert profiles["jm"].enabled is False


def test_daeseung_erp_credentials_use_payload_before_environment(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_VOUCHER_DAESEUNG_ERP_USER_ID", "env-user")
    monkeypatch.setenv("EXCEL_VOUCHER_DAESEUNG_ERP_PASSWORD", "env-password")
    corp_info = _apply_company_erp_credentials(
        {"company_key": "daeseung", "erp_credentials": {"user_id": "payload-user", "password": "payload-password"}},
        {"user_id": "old-user", "password": "old-password"},
    )

    assert corp_info["user_id"] == "payload-user"
    assert corp_info["password"] == "payload-password"


def test_erp_credentials_are_saved_per_user_and_company(tmp_path: Path) -> None:
    store = AccountStore(tmp_path / "accounts.sqlite3", initial_password="wowjd12!@")
    store.save_erp_credential("finance01", "daeseung", "12240413", "secret-pass")

    meta = store.get_erp_credential_meta("finance01", "daeseung")
    credential = store.get_erp_credential("finance01", "daeseung")

    assert meta and meta["erp_user_id"] == "12240413"
    assert credential == {"user_id": "12240413", "password": "secret-pass"}


def test_secret_protection_roundtrip() -> None:
    blob = protect_secret("secret-pass")

    assert "secret-pass" not in blob
    assert unprotect_secret(blob) == "secret-pass"


def test_resume_print_only_accepts_explicit_boolean_and_safe_string_flags(monkeypatch) -> None:
    monkeypatch.delenv("EXCEL_VOUCHER_PRINT_RECOVERY_ADMIN_IDS", raising=False)

    assert _resume_print_only_requested({"resume_print_only": True}) is True
    assert _resume_print_only_requested({"resume_print_only": "yes"}) is True
    assert _resume_print_only_requested({"resume_print_only": "0"}) is False
    assert _resume_print_only_requested({"resume_print_only": "false"}) is False


def test_resume_print_only_compatibility_signal_requires_admin_and_prefix(monkeypatch) -> None:
    monkeypatch.setenv("EXCEL_VOUCHER_PRINT_RECOVERY_ADMIN_IDS", "finance-admin, second-admin")
    payload = {
        "requester_id": "FINANCE-ADMIN",
        "source_filename": "__resume_print_only__voucher.xlsx",
    }

    assert _resume_print_only_requested(payload) is True
    assert _resume_print_only_requested({**payload, "requester_id": "ordinary-user"}) is False
    assert _resume_print_only_requested({**payload, "source_filename": "voucher.xlsx"}) is False


def test_resume_print_only_compatibility_signal_uses_default_admin(monkeypatch) -> None:
    monkeypatch.delenv("EXCEL_VOUCHER_PRINT_RECOVERY_ADMIN_IDS", raising=False)

    assert _resume_print_only_requested(
        {
            "requester_id": "rlckd9646",
            "source_filename": "__resume_print_only__6월 대승 수시결제.xlsx",
        }
    ) is True
